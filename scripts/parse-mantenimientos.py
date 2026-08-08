import openpyxl, json, re
from datetime import date

# ── Limpiar texto eliminando emojis y espacios extra ─────────────────────────
def limpiar(v):
    if v is None: return ''
    s = str(v).strip()
    # Eliminar emojis y caracteres no ASCII al inicio
    s = re.sub(r'^[\U00010000-\U0010ffff\u2600-\u26FF\u2700-\u27BF\uFE00-\uFE0F\s]+', '', s)
    return s.strip()

def norm(v): return limpiar(v).upper()

MESES_FULL = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
              'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']

def mes_idx(v):
    n = norm(v)
    if n in MESES_FULL: return MESES_FULL.index(n)
    for i,m in enumerate(MESES_FULL):
        if n and m.startswith(n[:3]): return i
    return -1

def fecha_iso(v):
    if not v: return ''
    if isinstance(v, date): return v.isoformat()
    if isinstance(v, str) and re.match(r'\d{2}/\d{2}/\d{4}', v):
        d,m,y = v.split('/')
        try: return date(int(y),int(m),int(d)).isoformat()
        except: return v
    return str(v).strip()

wb = openpyxl.load_workbook('/tmp/Mantenimientos_temp.xlsx')

# ── Mapa de celdas fusionadas ────────────────────────────────────────────────
def build_merge_map(ws):
    mm = {}
    for merged in ws.merged_cells:
        anchor_val = ws.cell(merged.min_row, merged.min_col).value
        for r in range(merged.min_row, merged.max_row+1):
            for c in range(merged.min_col, merged.max_col+1):
                if not (r==merged.min_row and c==merged.min_col):
                    mm[(r,c)] = anchor_val
    return mm

def get_cell(ws, merge_map, r, c):
    return merge_map.get((r,c), ws.cell(r,c).value)

# ── Parser PROGRAMADOS ────────────────────────────────────────────────────────
def parse_programados(ws):
    mm = build_merge_map(ws)

    # Encontrar fila de encabezados buscando la que tenga "MES" (limpiando emojis)
    hdr_row = -1
    for i in range(1, 6):
        row_vals = [norm(get_cell(ws, mm, i, c)) for c in range(1, ws.max_column+1)]
        if any('MES' == v for v in row_vals):
            hdr_row = i; break
    if hdr_row < 0:
        # Fallback: usar fila 3
        hdr_row = 3

    hdrs = [norm(get_cell(ws, mm, hdr_row, c)) for c in range(1, ws.max_column+1)]

    def colidx(patron):
        for i,h in enumerate(hdrs):
            if h and re.search(patron, h): return i+1
        return None

    iCopr  = colidx(r'COPROPIEDAD')
    iElem  = colidx(r'EQUIPO|ELEMENTO')
    iMes   = colidx(r'^MES$')
    iDia   = colidx(r'^D[IÍ]A$')
    iProv  = colidx(r'PROVEEDOR')
    iResp  = colidx(r'RESPONSABLE')
    iTipo  = colidx(r'TIPO')
    iObs   = colidx(r'OBSERVACI')
    iEst   = colidx(r'ESTADO')

    print(f'  Encabezados (fila {hdr_row}): Mes={iMes} Dia={iDia} Copr={iCopr} Elem={iElem} Est={iEst}')

    hoy = date.today(); anio = hoy.year
    programados = []

    for r in range(hdr_row+1, ws.max_row+1):
        def v(col): return get_cell(ws, mm, r, col) if col else None

        copr  = limpiar(v(iCopr))
        elem  = limpiar(v(iElem))
        mes_v = limpiar(v(iMes))
        dia_v = v(iDia)
        tipo  = limpiar(v(iTipo))  or 'Preventivo'
        estado= limpiar(v(iEst))  or 'Pendiente'
        prov  = limpiar(v(iProv)) or ''
        resp  = limpiar(v(iResp)) or ''
        obs   = limpiar(v(iObs))  or ''

        # Saltar filas vacías y separadores
        if not mes_v and not copr and not elem: continue
        if norm(mes_v) in ('MES', ''): continue

        mi = mes_idx(mes_v)
        dia_n = None
        try: dia_n = int(str(dia_v).strip()) if dia_v and str(dia_v).strip() else None
        except: pass

        prox = ''
        if mi >= 0 and dia_n:
            try:
                d = date(anio, mi+1, dia_n)
                if d < hoy: d = date(anio+1, mi+1, dia_n)
                prox = d.isoformat()
            except: pass

        mes_norm = 'MENSUAL' if norm(mes_v) == 'MENSUAL' else (MESES_FULL[mi] if mi>=0 else mes_v.upper())

        programados.append({
            '_copropiedad':     copr,
            '_elemento':        elem,
            'Copropiedad':      copr,
            'Elemento':         elem,
            'Mes':              mes_norm,
            'Dia':              str(dia_n) if dia_n else '',
            'Tipo':             tipo,
            'Proveedor':        prov,
            'Responsable':      resp,
            'Estado':           estado,
            'Observaciones':    obs,
            'Proximo_vencimiento': prox,
        })

    return programados

# ── Parser HISTORIAL ──────────────────────────────────────────────────────────
def parse_historial(ws):
    mm = build_merge_map(ws)
    hdr_row = -1
    for i in range(1,5):
        row_vals = [norm(get_cell(ws,mm,i,c)) for c in range(1,ws.max_column+1)]
        if any('FECHA' in v for v in row_vals if v):
            hdr_row=i; break
    if hdr_row<0: hdr_row=3

    hdrs = [norm(get_cell(ws,mm,hdr_row,c)) for c in range(1,ws.max_column+1)]
    def colidx(p):
        for i,h in enumerate(hdrs):
            if h and re.search(p,h): return i+1
        return None

    iCopr  = colidx(r'COPROPIEDAD')
    iElem  = colidx(r'EQUIPO|ELEMENTO')
    iTipo  = colidx(r'TIPO')
    iFecha = colidx(r'FECHA')
    iProv  = colidx(r'PROVEEDOR')
    iResp  = colidx(r'RESPONSABLE')
    iRes   = colidx(r'RESULTADO')
    iObs   = colidx(r'OBSERVACI')

    historial = []
    for r in range(hdr_row+1, ws.max_row+1):
        def v(col): return get_cell(ws,mm,r,col) if col else None
        copr  = limpiar(v(iCopr))
        elem  = limpiar(v(iElem))
        fecha = fecha_iso(v(iFecha))
        if not fecha and not copr and not elem: continue
        historial.append({
            '_copropiedad': copr,
            '_elemento':    elem,
            'Copropiedad':  copr,
            'Elemento':     elem,
            'Tipo':         limpiar(v(iTipo)),
            'Fecha_ejecucion': fecha,
            'Proveedor':    limpiar(v(iProv)),
            'Responsable':  limpiar(v(iResp)),
            'Resultado':    limpiar(v(iRes)),
            'Observaciones':limpiar(v(iObs)),
        })
    return historial

ws_prog = wb['PROGRAMADOS']
ws_hist = wb['HISTORIAL']
prog = parse_programados(ws_prog)
hist = parse_historial(ws_hist)

out = {
    'actualizado': date.today().isoformat(),
    'totalProgramados': len(prog),
    'totalHistorial':   len(hist),
    'programados': prog,
    'historial':   hist,
}
with open('mantenimientos.json','w') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f'OK: {len(prog)} programados, {len(hist)} historial')
